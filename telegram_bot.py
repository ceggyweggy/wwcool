import requests

from config import TELEGRAM_SEND_MESSAGE_URL

import time 
import random 
import telegram_bot 
import telegram 
import telegram.ext 
import math 
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from s_80fenAI import s_tractor_AI 

random.seed()

# intro - 12/12/2020, updated 16/12/2020
intro_names = ["yuxuan", "zanna", "andreaaaaaa", "emma", "adele", "chloe", "jennifer", "peixin", "snow", "xaviera", "fei", "cadence wern", "raewyn"]
intro_classes = ["312", "304", "312!!!", "308", "204", "204/313", "308", "302.", "302", "308", "this bot sucks", "307", "305"]
intro_cca = ["chinese orchestra", "drama", "girl guidesssssss wuhu", "robo!!", "shooting", "robo", "jingle bells", "ncc.", "art club", "red cross", "suicide squad", "robo", "red cross"]
intro_id = [1235169098, 1149401183, 1254075693, 1206367870, 1119686699, 712481525, 1045960315, 1263449655, 1196893557, 1123398215, 771649055, 861886075, 811427963]
intro_subj = ['3ScH', 'trip lit','trip hist yay :))', 'cphl', 'nil', 'trip hist,', 'pchl', 'nil', "trip geog", "cphl", "trip hist", "pcgh!!", "trip hist"]
intro_game = ['quizarium', 'WW', 'werewolf cuz idk it\'s fun hehe but bridge close second yay', 'everything but bridge cause it\'s too big brain for me',
'nil', 'a l l (or ma jiang)', 'nil', 'nil', "all tbh", "ww", "ww", "bridge ha andrea", "wordchain"]

# bridge stats (new version) - 11/06/2021
bridge_stats = load_workbook('bridgescores.xlsx')
standings = bridge_stats['Standings']
raw_scores = bridge_stats['Raw Scores']
def update_bridge(name1, name2):
    global bridge_stats, standings, raw_scores
    number_of_entries = raw_scores['E1']
    new_entry = 'A' + str(number_of_entries.value+2)
    raw_scores[new_entry].value = datetime.today()
    new_entry = 'B' + str(number_of_entries.value+2)
    raw_scores[new_entry].value = name1
    new_entry = 'C' + str(number_of_entries.value+2)
    raw_scores[new_entry].value = name2
    number_of_entries.value += 1

    number_of_winners = standings['E1']
    one_found = False
    two_found = False
    for i in range(2, number_of_winners.value+2):
        cell = 'A' + str(i)
        name = standings[cell].value
        if name == name1 or name == name2:
            if name == name1:
                one_found = True
            else:
                two_found = True
            cell = 'B' + str(i)
            standings[cell].value += 1
    if not one_found:
        number_of_winners.value += 1
        cell = 'A' + str(number_of_winners.value)
        standings[cell].value = name1
        cell = 'B' + str(number_of_winners.value)
        standings[cell].value = 1 
    if not two_found:
        number_of_winners.value += 1
        cell = 'A' + str(number_of_winners.value)
        standings[cell].value = name2
        cell = 'B' + str(number_of_winners.value)
        standings[cell].value = 1
    bridge_stats.save('bridgescores.xlsx')
    # print(printscores())

# tictactoe update - 28/12/2020
ttt = []
starter = []
opponent = []
turn = []

# control bot update - admin actions - 14/12/2020
admin_id = 861886075
send_id = 0
msg_id = 0 

# fixed bug - 26/12/2020, updated 10/1/2021
unixtime = int(time.time())
current_id = 975589315
times = 0 

# send in orig form - 3/1/2021
orig_text = '' 

# s chain hints - 9/3/2021
words = ["sacrifices", "sacrilegious", "sagittarius", "salamanders", "saltinesses", "salubrious", "salubriousness", "salutations", "sandstorms", "satisfactions", "saturations", "saxophones", "scandalous", "scarecrows", "scatterings", "schoolboys", "schoolgirls", "schoolmasters", "schoolmates", "scoreboards", "scornfulness", "screwdrivers", "seasickness", "seasicknesses", "securities", "sedimentations", "seductiveness", "segregatedness", "segregations", "seldomness", "selfishness", "selflessness", "semicircles", "semiconductors", "semiconscious", "sensitivities", "sentiments", "seriousness", "severeness", "sexagenarians", "sextuplets", "shamefulness", "sharpeners", "shellfishes", "shipwrecks", "shittiness", "shrubberies", "simpleness", "simpletons", "simplisms", "simultaneous", "skateboarders", "skateboards", "skepticalness", "skyrockets", "sloppiness", "smartnesses", "somebodies", "somersaults", "specialists", "specialness", "specialties", "specifications", "specificness", "spectacles", "spontaneous", "spontaneousness", "sportsmanships", "squeakiness", "squeamishness", "stammeringness", "starboards", "starlights", "startlingness", "starvations", "statistics", "steppingstones", "stereotypes", "straightforwardness", "strategicness", "strategies", "strawberries", "stupidness", "subconscious", "subconsciousness", "submarines", "subordinates", "subscribers", "subscripts", "sugariness", "supercalifragilisticexpialidocious", "superconductors", "superheroes", "superheros", "superintendents", "superlatives", "supermarkets", "superordinates", "superstitions", "superstitious", "supplementaries", "supplements", "suppressiveness", "surreptitiousness", "surroundings", "swampiness", "sweatiness", "sweetenings", "sweetnesses", "swordfishes", "syllogisms", "sympatheticness", "symptomaticness", "synchronous"]
to_approve = [] 

# 80fen part 1 - 23/5/2021
tractor_ids = [] # to fill with id 
tractor_games = [] # to fill w games themselves 
tractor_selfids = [] # to fill with player ids

# 80fen scores tracking
tractor_scores = load_workbook('80fenscores.xlsx')
tractor_standings = tractor_scores['Standings']
tractor_rawscores = tractor_scores['Raw Scores']

def Diff(li1, li2):
    li_dif = [i for i in li1 + li2 if i not in li1 or i not in li2]
    return li_dif

def name_to_id(name):
    if name=="ww cool":
        chat_id=-1001199978565
    elif name.startswith("cadence") or name=="cadonk" or name.startswith("cegg"):
        chat_id=861886075
    elif name.startswith("chloe") or name.startswith("crowee") or name.startswith("cloe"):
        chat_id=712481525
    elif name=="xf" or name=="xuefei":
        chat_id=771649055
    elif name.startswith("adele") or name.startswith("defenestration") :
        chat_id=1119686699
    elif name=="yx" or name=="yuxuan":
        chat_id=1235169098
    elif name=="emma":
        chat_id=1206367870
    elif name=="zanna" or name=="zyles" or name=="zxnna":
        chat_id=1149401183
    elif name.startswith("past curfew"):
        chat_id=-434170116
    elif name=="teapot" or name=="raewyn":
        chat_id=811427963
    elif name.startswith("xavii") or name.startswith("xaviera"):
        chat_id=1123398215
    elif name=="hihihi":
        chat_id=-452778702
    elif name=="celina":
        chat_id=877122271
    elif name=="safer":
        chat_id=-1001219366779
    elif name.startswith('jen\'s mom'):
        chat_id = 1827751461
    elif name.startswith('jen') or name.startswith('hen'):
        chat_id=1045960315
    elif name.startswith('andrea') or name.startswith('nadre'):
        chat_id=1254075693
    elif name.startswith('pect') or name.startswith('peixin'):
        chat_id=1263449655
    elif name.startswith('snow'):
        chat_id=1196893557
    elif name.startswith('sonya') or name.startswith('bitch'):
        chat_id=918254418
    elif name.startswith('bensy') or name.startswith('benita'):
        chat_id=1119831098
    elif name.startswith('gab'):
        chat_id=1150432357
    elif name.startswith('ruike') or name.startswith('rk'):
        chat_id=1740633932
    elif name.startswith('dan'):
        chat_id=1063667057
    elif name.startswith('kang ying') or name.startswith('ky') or name.startswith('kangying'):
        chat_id=1140148199
    elif name.startswith('80fen'):
        chat_id=-1001498506051
    elif name.startswith('lanche') or name.startswith('averlynn'):
        chat_id=1343835043
    elif name.startswith('clavin'):
        chat_id=1552784611
    else:
        chat_id=0
    return chat_id

def id_to_name(chat_id):
    if chat_id==-1001199978565:
        username='ww cool'
    elif chat_id==861886075:
        username='Cadence'
    elif chat_id==712481525:
        username='Chloe'
    elif chat_id==771649055:
        username='Xuefei'
    elif chat_id==1119686699:
        username='Adele'
    elif chat_id==1235169098:
        username='Yuxuan'
    elif chat_id==1206367870:
        username='Emma'
    elif chat_id==1149401183:
        username='Zanna'
    elif chat_id==-434170116:
        username='Past Curfew WW Cool'
    elif chat_id==811427963:
        username="Raewyn"
    elif chat_id==1123398215:
        username="Xaviera"
    elif chat_id==-452778702:
        username="hihihi"
    elif chat_id==877122271:
        username="Celina"
    elif chat_id==-1001219366779:
        username='Safer'
    elif chat_id==1045960315 or chat_id==1827751461:
        username='Jennifer'
    elif chat_id==1254075693:
        username='Andrea'
    elif chat_id==1263449655:
        username='Peixin'
    elif chat_id==1196893557:
        username='Snow'
    elif chat_id==918254418:
        username='Sonya'
    elif chat_id==1119831098:
        username='Benita'
    elif chat_id==1150432357:
        username='Gabriel'
    elif chat_id==1740633932:
        username='Ruike'
    elif chat_id==1063667057:
        username='Dan'
    elif chat_id==1140148199:
        username='Kang Ying'
    elif chat_id==-1001498506051:
        username="80fen"
    elif chat_id==1343835043:
        username="Averlynn"
    elif chat_id==1552784611:
        username="Clavin"
    elif chat_id==0:
        username="Bot0"
    elif chat_id==1:
        username="Bot1"
    elif chat_id==2:
        username="Bot2"
    else:
        username=''
    return username

# 1/6/2021 - DONE WITH ALL EXCEPT ADJ PAIRS!!!
class tractor():
    def start(self):
        # starting 
        self.player_names = []
        self.is_bot = [0, 0, 0, 0]
        self.bot_objs = [None, None, None, None]

        # print funcs
        self.message = '' 
        self.personal_message = ''

        # dealing cards 
        self.suits = ['c', 'd', 'h', 's']
        self.suit_names = ['clubs', 'diamonds', 'hearts', 'spades']
        self.deck_nums = ['2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K', 'A', 'X'] # joker 
        self.deck = [] 
        self.draw_counter = 0
        self.deck_counter = 0
        self.player_decks = {0:[], 1:[], 2:[], 3:[]}

        # decide tao fen 
        self.tao_fen = True
        self.tao_fen_card = ''
        self.tao_fen_person = 3

        # decide trump suit 
        self.decided_trump = False
        self.trump_suit = ''
        self.trump_value = 'J'
        self.other_trump = '3'
        self.one_trumps_played = False
        self.two_trumps_played = False 

        # making 底
        self.making_di = False
        self.di = []
        self.di_points = 0

        # confirming stuff 
        self.to_confirm = False
        self.confirming = -1
        
        # playing 
        self.actually_playing = False 
        self.moving_player = -1
        self.first_player = True
        self.first_player_id = 3
        self.player_moves = {0:[], 1:[], 2:[], 3:[]}
        self.mode = 0
        self.move_counter = 0
        self.aces_suits = []
        self.current_winner = -1
        self.same_pattern = False 
        self.cards_left = 0

        # points counting
        self.qiang_fen_points = 0 
        self.points = 0

        self.game_ended = False 

        # starting stuff
        self.make_deck()
        self.message = "Started successfully! Use /join80fen to join the game :cD"
        return

    def get_suit(self, card):
        if self.is_trump(card):
            return self.trump_suit
        return card[0]

    def add_bot(self):
        self.is_bot[len(self.player_names)] = 1
        self.bot_objs[len(self.player_names)] = s_tractor_AI(len(self.player_names))
        self.player_names.append(str(len(self.player_names)))
        if len(self.player_names) == 4:
            self.auto_deal()

    def auto_deal(self):
        for i in range(100):
            self.player_decks[i%4].append(self.deck[i])
        self.deck_counter = 100 
        self.making_di = True 
        self.message = "Done dealing!"

        for i in range(4):
            if self.is_bot[i]:
                self.bot_objs[i].deck = self.player_decks[i]
                trumpy = self.bot_objs[i].call_trump()
                if trumpy != '':
                    print(trumpy)
                    if self.bot_objs[i].deck.count(trumpy) > 1:
                        self.trump_suit = trumpy[0]
                        self.decided_trump = True
                        self.message = id_to_name(self.player_names[i]) + " has played two copies of the trump suit they want! The trump suit is " + self.suit_names[self.suits.index(self.trump_suit)] + " :c)"
                        return self.message
                        # self.check_bot_di()
                        
                    else:
                        self.one_trumps_played = True
                        self.trump_suit = trumpy[0]
                        self.message = id_to_name(self.player_names[i]) + " has played one copy of the card they wish to see as trump suit! The trump suit is now " + self.suit_names[self.suits.index(self.trump_suit)] + ". To change the trump suit, you must play two copies of " + str(self.trump_value)
                        return self.message


    def join(self, name):
        self.player_names.append(name)
        self.message = id_to_name(name) + " has been added to the game! "
        if len(self.player_names) == 4:
            self.auto_deal()
            # self.message += id_to_name(self.player_names[0]) + ", draw a card! (/drawcard)"

    def make_deck(self):
        for i in range(13):
            for j in range(4):
                card = self.suits[j] + self.deck_nums[i]
                self.deck.insert(0, card)
                self.deck.insert(0, card)
        self.deck.insert(0, "rX")
        self.deck.insert(0, "rX")
        self.deck.insert(0, "bX")
        self.deck.insert(0, "bX")
        random.shuffle(self.deck)
        # for i in range(len(self.deck)):
        #   print(self.deck[i])

    def draw(self, name):
        if not name == self.player_names[self.draw_counter]:
            self.message = "Wrong person drawing! " + self.player_names[self.draw_counter] + " is supposed to draw a card now."
            return self.message

        if self.deck_counter == 100:
            self.message = "All cards drawn!"
            return self.message

        self.player_decks[self.draw_counter].insert(0, self.deck[self.deck_counter])
        self.print_deck(self.player_names[self.draw_counter])

        if self.trump_value in self.deck[self.deck_counter] and not self.tao_fen:
            self.personal_message += "\nYou can play " +  " to be 逃分! Use /<card> to play"

        self.deck_counter += 1
        self.draw_counter += 1
        self.draw_counter %= 4

        if self.deck_counter == 100:
            for i in range(4):
                self.player_decks[i].sort()
            self.message = "All cards drawn!\n"
            self.message += id_to_name(self.player_names[0])
            self.message += ", /drawdi to draw your 底 :cD"
            self.making_di = True 
            return self.personal_message

        self.message = id_to_name(self.player_names[self.draw_counter])
        self.message += ", draw a card! (/drawcard)"
        return self.personal_message

    def print_deck(self, name):
        if not name in self.player_names:
            self.message = "You're not playing the game!"
            return 
        self.personal_message = ''
        index = self.player_names.index(name)
        self.player_decks[index].sort()
        for i in range(len(self.player_decks[index])):
            if i == 0:
                self.personal_message += "/"
            elif self.player_decks[index][i][0] == self.player_decks[index][i-1][0]:
                self.personal_message += ", /"
            else:
                self.personal_message += "\n/"
            self.personal_message += self.player_decks[index][i]
        return 

    def draw_di(self, name):
        if not name == self.player_names[self.tao_fen_person]:
            self.message = "You can't do that."
            return self.message
        if self.trump_suit == '':
            self.message = "Decide a trump first."
            return self.message
        if not self.decided_trump:
            self.decided_trump = True 
        self.making_di = True 
        for i in range(100, 108):
            self.player_decks[self.tao_fen_person].append(self.deck[i])
        self.print_deck(name)
        self.personal_message += "\n Done! To select cards to place in your 底, use /<card>, eg: /cA. \nRemember to include 8 cards in your 底 :c)"
        return self.personal_message

    def check_tao_fen(self, card, name):
        if not self.trump_value in card:
            return
        self.personal_message = "pls /confirm thx (this is for 逃分 btw)"
        self.to_confirm = True 
        self.tao_fen_card = card
        self.confirming = self.player_names.index(name)
        return self.personal_message

    def check_making_di(self, card, name):
        # print(card) 
        if not name == self.player_names[self.tao_fen_person]:
            return 
        if card in self.player_decks[self.tao_fen_person]:
            # move to 底
            self.di.append(card)
            self.player_decks[self.tao_fen_person].remove(card)

        elif card in self.di:
            # remove from 底
            while card in self.di:
                self.player_decks[self.tao_fen_person].append(card)
                self.di.remove(card)

        else:
            return ''

        self.print_deck(self.player_names[self.tao_fen_person])
        self.cards_left = 8 - len(self.di)
        self.personal_message += "\nYour 底: "
        for i in range(len(self.di)):
            if not i == 0:
                self.personal_message += ", "
            self.personal_message += "/" + self.di[i] 

        self.personal_message += "\nSuccess!\n"

        if self.cards_left == 0:
            self.personal_message = "/confirm to confirm your 底 or you can /<card> to change your 底"
            self.to_confirm = True
            self.confirming = self.tao_fen_person
        return self.personal_message

    def check_trump_suit(self, card, name):
        if not self.trump_value in card:
            return ''
        number = self.player_decks[self.player_names.index(name)].count(card)
        if number == 1 and not self.one_trumps_played:
            self.trump_suit = card[0]
            index = self.suits.index(self.trump_suit)
            self.personal_message = "pls /confirm thx (this is for trump btw)"
            self.to_confirm = True
            self.one_trumps_played = True 
            self.confirming = self.player_names.index(name)
        if number == 2:
            self.trump_suit = card[0]
            index = self.suits.index(self.trump_suit)
            self.personal_message = "pls /confirm thx (this is for two trumps btw)"
            self.one_trumps_played = False 
            self.two_trumps_played = True 
            self.to_confirm = True 
            self.confirming = self.player_names.index(name)
        return self.personal_message

    def check_playing_cards(self, card, name):
        index = self.player_names.index(name)
        print(index)
        print(card)
        if card in self.player_decks[index]:
            # move to moving pile
            self.player_moves[index].append(card)
            self.player_decks[index].remove(card)
        elif card in self.player_moves[index]:
            # move to deck
            while card in self.player_moves[index]:
                self.player_moves[index].remove(card)
                self.player_decks[index].append(card)
        else:
            return ''
        self.print_deck(self.player_names[self.moving_player])
        self.personal_message += "\n"
        self.personal_message += "Cards you wish to play: "
        for i in range(len(self.player_moves[index])):
            if not i == 0:
                self.personal_message += ", "
            self.personal_message += "/" + self.player_moves[index][i]
        if self.first_player:
            self.personal_message += "\nYou can play more cards, or you can /confirm to confirm your move."
            self.to_confirm = True
            self.confirming = self.player_names.index(name)
            return self.personal_message
        self.cards_left = len(self.player_moves[self.first_player_id]) - len(self.player_moves[index])
        if self.cards_left:
            self.personal_message += "\nYou have to play " + str(self.cards_left) + " more cards."
            self.to_confirm = False
            self.confirming = -1
            return self.personal_message
        self.personal_message += "\nYou can /confirm to confirm your move, or /<card> to change the cards you wish to play."
        self.to_confirm = True
        self.confirming = self.player_names.index(name)
        return self.personal_message

    def whats_trump(self):
        if self.trump_suit == '':
            self.message = "There is no trump yet. The trump value is " + self.trump_value
        elif self.decided_trump:
            self.message = "The trump suit is " + self.suit_names[self.suits.index(self.trump_suit)] + ", and it can't be changed. The trump value is " + self.trump_value
        else:
            self.message = "The trump suit is currently " + self.suit_names[self.suits.index(self.trump_suit)] + ", and it can be changed. The trump value is " + self.trump_value
        return self.message

    def print_moves(self):
        self.message = '' 
        for i in range(4):
            self.player_moves[i].sort()
            self.message += id_to_name(self.player_names[i]) + " - "
            for j, card_played in enumerate(self.player_moves[i]):
                if j != 0:
                    self.message += ", "
                self.message += card_played
            self.message += '\n'
        return self.message

    def check_valid_cards(self):
        if self.cards_left:
            self.personal_message = "Wrong number of cards played!"
            return False

        if not len(self.player_moves[self.moving_player]):
            # no cards played 😂
            self.personal_message = "You didn't play any cards 😂"
            return False 

        # first player 
        if self.mode == 0:
            self.mode = self.set_mode()

        try:
            # 1 card - DONE 
            if self.mode == 1:
                # print(self.player_moves[self.moving_player][0])
                # one card played
                # first player can play anyth 
                if self.first_player:
                    self.current_winner = self.first_player_id
                    return True 
                if self.is_trump(self.player_moves[self.first_player_id][0]):
                    card_suit = self.trump_suit
                else:
                    card_suit = self.player_moves[self.first_player_id][0][0]
                if self.is_trump(self.player_moves[self.moving_player][0]):
                    mover_card_suit = self.trump_suit
                else:
                    mover_card_suit = self.player_moves[self.moving_player][0][0]

                if card_suit == mover_card_suit:
                    # same suit 
                    return True 

                # not the same suit, check for if cards of same suit 
                for i, deck_card in enumerate(self.player_decks[self.moving_player]):
                    if card_suit != self.trump_suit:
                        if card_suit == deck_card[0] and not self.is_trump(deck_card):
                            self.personal_message = "You still have cards of the suit (?) (not trump?)"
                            return False
                    elif self.is_trump(deck_card):
                        self.personal_message = "You still have cards of the suit (?) (trump)"
                        return False 
                return True 

            # pair - DONE
            elif self.mode == 2:
                if self.is_trump(self.player_moves[self.first_player_id][0]):
                    card_suit = self.trump_suit
                else:
                    card_suit = self.player_moves[self.first_player_id][0][0]
                if self.is_trump(self.player_moves[self.moving_player][0]):
                    mover_card_suit = self.trump_suit
                else:
                    mover_card_suit = self.player_moves[self.moving_player][0][0]

                # same suit + pair
                if self.player_moves[self.moving_player][0] == self.player_moves[self.moving_player][1]:
                    self.same_pattern = True 
                    if card_suit == mover_card_suit:
                        same_suit = True 
                        return True
                else:
                    self.same_pattern = False
                if self.first_player:
                    self.personal_message = "Invalid sequence (detected as: pair)"
                    return False
                else:
                    same_suit = True
                    for i, card_played in enumerate(self.player_moves[self.moving_player]):
                        if self.is_trump(card_played) and card_suit != self.trump_suit:
                            same_suit = False
                            break
                        elif card_suit == self.trump_suit and not self.is_trump(card_played):
                            same_suit = False
                            break
                    if same_suit:
                        # check if have pair 
                        for i, deck_card in enumerate(self.player_decks[self.moving_player]):
                            if card_suit == self.trump_suit:
                                if not self.is_trump(deck_card):
                                    continue
                            elif self.is_trump(deck_card) or deck_card[0] != card_suit:
                                continue
                            if self.player_decks[self.moving_player].count(deck_card) + self.player_moves[self.moving_player].count(deck_card) > 1:
                                self.personal_message = "You still have a pair! (" + deck_card + ")"
                                return False
                    else:
                        # suit changed 
                        for i, deck_card in enumerate(self.player_decks[self.moving_player]):
                            if card_suit == self.trump_suit and self.is_trump(deck_card):
                                self.personal_message = "You still have cards of the suit! (" + deck_card + ')'
                                return False
                            elif not self.is_trump(deck_card) and card_suit == deck_card[0]:
                                self.personal_message = "You still have cards of the suit! (" + deck_card + ')'
                                return False
                    return True 

            # adj pair - TO TEST 
            elif self.mode == 3:
                # adj pairs 
                # find card suit 
                if self.is_trump(self.player_moves[self.first_player_id][0][0]):
                    card_suit = self.trump_suit
                else:
                    card_suit = self.player_moves[self.first_player_id][0][0]
                # check if adj pairs also played
                adj_pairs_played = True 
                card_vals = [] 
                for i, deck_card in enumerate(self.player_moves[self.moving_player]):
                    if i%2 == 1:
                        card_vals.append(deck_card[1:])
                        continue
                    if deck_card == self.player_moves[self.moving_player][i+1]:
                        continue
                    else:
                        adj_pairs_played = False
                        break
                card_vals.sort()
                for i in range(1, len(card_vals)):
                    if self.deck_nums.index(card_vals[i]) - self.deck_nums.index(card_vals[i-1]) > 1:
                        adj_pairs_played = False 
                        break
                if adj_pairs_played:
                    self.same_pattern = True
                    return True
                self.same_pattern = False
                # first player must play adj pairs
                if self.first_player:
                    self.personal_message = "Invalid first player sequence! (detected as: adjacent pairs)"
                    return False 
                # if not: check for adj pairs
                # check for pairs in deck and in move played --> adjacent?
                pairs = [] 
                same_suit_played = True 
                cards_of_suit_left = False
                pairs_played = 0 
                for i, player_card in enumerate(self.player_decks[self.moving_player]):
                    if card_suit != self.trump_suit:
                        if not self.is_trump(player_card) and player_card[0] == card_suit:
                            cards_of_suit_left = True
                            # same suit, check if its a pair 
                            if self.player_decks[self.moving_player].count(player_card) + self.player_moves[self.moving_player].count(player_card) > 1:
                                # yes, append to pairs
                                pairs.append(player_card)
                    # trumps played 
                    elif self.is_trump(player_card):
                        cards_of_suit_left = True
                        # same suit, check if its a pair 
                        if self.player_decks[self.moving_player].count(player_card) + self.player_moves[self.moving_player].count(player_card) > 1:
                            # yes, append to pairs
                            pairs.append(player_card)

                for i, player_card in enumerate(self.player_moves[self.moving_player]):
                    if not self.is_trump(player_card) and card_suit == player_card[0]:
                        if self.player_decks[self.moving_player].count(player_card) + self.player_moves[self.moving_player].count(player_card) > 1 and player_card not in pairs:
                            pairs_played += 1
                            # yes, append to pairs
                            pairs.append(player_card)
                    elif self.is_trump(player_card) and card_suit == self.trump_suit:
                        if self.player_decks[self.moving_player].count(player_card) + self.player_moves[self.moving_player].count(player_card) > 1 and player_card not in pairs:
                            pairs_played += 1
                            # yes, append to pairs
                            pairs.append(player_card)
                    else:
                        same_suit_played = False

                # pairs: convert to value
                for i, pair in enumerate(pairs):
                    pairs[i] = self.deck_nums.index(pair[1:])


                pairs.sort()

                for i in range(1, len(pairs)):
                    if pairs[i] - pairs[i-1] == 1:
                        self.personal_message = "Adj pairs exist in ur deck lol"
                        return False

                # then check for pairs
                if pairs_played == len(self.player_moves[self.first_player_id])/2:
                    return True
                # elif len(pairs) < pairs_played:
                #     self.personal_message = "Unknown Error. (Number of calculated pairs < number of actual pairs??)"
                #     return False 
                if len(pairs) > pairs_played:
                    self.personal_message = "You didn't play all of your pairs."
                    return False 

                # then check for suit 
                if same_suit_played:
                    return True
                if cards_of_suit_left:
                    self.personal_message = "You still have cards of the suit! (" + self.suit_names[self.suits.index(card_suit)] + ') Play those first. Your move sequence should already be cleared, so /showdeck to see your deck again or /<card> to make your move.'
                    return False
                if self.get_trump_hierarchy(self.player_moves[self.moving_player][0]) and adj_pairs_played:
                    self.same_pattern = True
                return True

            # aces only - TO TEST
            elif self.mode == 4:
                # all aces 
                player_suits = [] 
                if self.first_player:
                    if self.move_counter:
                        self.personal_message = "Only can be played on first move!"
                        return False
                    for i, card_played in enumerate(self.player_moves[self.first_player_id]):
                        if self.is_trump(card_played):
                            self.aces_suits.append(self.trump_suit)
                        else:
                            self.aces_suits.append(card_played[0])
                    self.aces_suits.sort()
                    return True 
                # not first player
                for i, card_played in enumerate(self.player_moves[self.moving_player]):
                    if self.is_trump(card_played):
                        player_suits.append(self.trump_suit)
                    else:
                        player_suits.append(card_played[0])
                player_suits.sort()
                # check if suits r the same 
                if player_suits == self.aces_suits:
                    self.same_pattern = True 
                    return True 
                self.same_pattern = False 
                difference = Diff(self.aces_suits, player_suits)
                for i, diff_suit in enumerate(difference):
                    if diff_suit in self.aces_suits:
                        for j, deck_card in enumerate(self.player_decks[self.moving_player]):
                            if diff_suit in deck_card:
                                self.personal_message = "You still have cards of the suit! (" + deck_card + ') Your move sequence should already be cleared, so /showdeck to see your deck again or /<card> to make your move.'
                                return False 
                return True

        except IndexError as e:
            raise IndexError from e
            return True

    def is_trump(self, card):
        if self.trump_suit in card:
            return True 
        if self.trump_value in card:
            return True
        if 'r' in card or 'b' in card:
            return True 
        return False 

    def set_mode(self):
        # the check is separate from the set_mode 
        if len(self.player_moves[self.moving_player]) == 1:
            return 1
        elif len(self.player_moves[self.moving_player]) == 2 and self.player_moves[self.moving_player][0] == self.player_moves[self.moving_player][1]:
            return 2
        for i, card_played in enumerate(self.player_moves[self.moving_player]):
            if not 'A' in card_played:
                return 3
        return 4

    def clear_moves(self):
        for i in range(4):
            self.player_moves[i].clear()

    def get_trump_hierarchy(self, card):
        if self.trump_suit in card and self.trump_value not in card:
            return 1
        if self.trump_suit not in card and self.trump_value in card:
            return 2
        if self.trump_suit in card and self.trump_value in card:
            return 3
        if 'b' in card:
            return 4
        if 'r' in card:
            return 5
        return 0 

    def decide_winner(self):
        if self.first_player:
            self.current_winner = self.first_player_id
            return

        if self.mode == 1:
            a = self.get_trump_hierarchy(self.player_moves[self.current_winner][0])
            b = self.get_trump_hierarchy(self.player_moves[self.moving_player][0])
            if b > a:
                self.current_winner = self.moving_player
            elif a == b == 0 or a == b == 1:
                if self.player_moves[self.current_winner][0][0] != self.player_moves[self.moving_player][0][0]:
                    return 
                current_highest_value = self.deck_nums.index(self.player_moves[self.current_winner][0][1:])
                new_card_value = self.deck_nums.index(self.player_moves[self.moving_player][0][1:])
                if current_highest_value < new_card_value:
                    self.current_winner = self.moving_player

        if self.mode == 2:
            if not self.same_pattern:
                return 
            a = self.get_trump_hierarchy(self.player_moves[self.current_winner][0])
            b = self.get_trump_hierarchy(self.player_moves[self.moving_player][0])
            if b > a:
                self.current_winner = self.moving_player
            elif a == b == 0 or a == b == 1:
                if self.player_moves[self.current_winner][0][0] != self.player_moves[self.moving_player][0][0]:
                    return
                current_highest_value = self.deck_nums.index(self.player_moves[self.current_winner][0][1:])
                new_card_value = self.deck_nums.index(self.player_moves[self.moving_player][0][1:])
                if current_highest_value < new_card_value:
                    self.current_winner = self.moving_player

        if self.mode == 3:
            if not self.same_pattern:
                return 
            a = self.get_trump_hierarchy(self.player_moves[self.current_winner][0])
            b = self.get_trump_hierarchy(self.player_moves[self.moving_player][0])
            if b > a:
                self.current_winner = self.moving_player
            elif a == b == 0 or a == b == 1:
                if self.player_moves[self.current_winner][0][0] != self.player_moves[self.moving_player][0][0]:
                    return
                self.current_highest_value = self.deck_nums.index(self.player_moves[self.current_winner][0][1:])
                self.new_card_value = self.deck_nums.index(self.player_moves[self.moving_player][0][1:])
                if self.deck_nums[self.deck_nums.index(self.current_highest_value)] < self.deck_nums[self.deck_nums.index(self.new_card_value)]:
                    self.current_winner = self.moving_player

        if self.mode == 4:
            if self.same_pattern:
                return 
            for i in range(len(self.player_moves[self.moving_player])):
                if (self.trump_suit not in self.player_moves[self.moving_player][i]) or (self.trump_value not in self.player_moves[self.moving_player][i]) or ('r' not in self.player_moves[self.moving_player][i]) or ('b' not in self.player_moves[self.moving_player][i]):
                    return
            self.current_winner = self.moving_player
            # cmon js give them the credit yknow ?? 

    def add_points(self):
        for i in range(len(self.player_moves[self.moving_player])):
            if '5' in self.player_moves[self.moving_player][i]:
                self.points += 5
            elif '10' in self.player_moves[self.moving_player][i] or 'K' in self.player_moves[self.moving_player][i]:
                self.points += 10

    def confirm(self, name):
        if not self.to_confirm:
            self.message = "There is nothing to confirm."
            return self.message

        if not self.confirming == self.player_names.index(name):
            self.personal_message = "You don't have anything to confirm."
            return self.personal_message

        # tao fen decided
        if not self.tao_fen:
            print("Tao fen!")
            self.message = id_to_name(name) + " has played " + self.tao_fen_card + "! They will now play 逃分 with their partner (" + id_to_name(self.player_names[(self.player_names.index(name)+2)%4]) + "). The other two will play 抢分 :c). Also this means that " + id_to_name(name) + " will draw 底 (jen i coded this for u appreciate me rn)"
            self.tao_fen = True
            self.to_confirm = False
            self.confirming = -1 
            self.tao_fen_person = self.player_names.index(name)
            self.first_player_id = self.tao_fen_person
            return self.message

        # 1 trump decided
        elif not self.decided_trump and self.one_trumps_played:
            print("One trump!")
            index = self.suits.index(self.trump_suit)
            self.message = id_to_name(name) + " has played one copy of the card they wish to see as trump suit! The trump suit is now " + self.suit_names[index] + ". To change the trump suit, you must play two copies of " + str(self.trump_value)
            self.to_confirm = False
            self.confirming = -1
            return self.message

        # 2 trumps decided (final)
        elif not self.decided_trump and self.two_trumps_played:
            print("Two trumps!")
            index = self.suits.index(self.trump_suit)
            self.message = id_to_name(name) + " has played two copies of the trump suit they want! The trump suit is " + self.suit_names[index] + " :c)"
            self.decided_trump = True 
            self.to_confirm = False
            self.confirming = -1
            return self.message

        # making 底
        elif self.making_di and name == self.player_names[self.tao_fen_person]:
            print("Making di!")
            for i, di_card in enumerate(self.di):
                if '5' in di_card:
                    self.di_points += 5
                elif '10' in di_card or 'K' in di_card:
                    self.di_points += 10
            self.di_points *= 2
            self.making_di = False
            self.actually_playing = True 
            self.moving_player = self.tao_fen_person
            self.message = id_to_name(name) + " has finally finished making their 底, and can now go on to play some cards." 
            self.print_deck(self.player_names[self.moving_player])
            self.to_confirm = False
            self.confirming = -1 
            return self.message

        # playing cards 
        elif self.actually_playing and self.moving_player == self.player_names.index(name):
            print("Playing cards!")
            self.player_moves[self.moving_player].sort()
            if self.check_valid_cards():
                self.print_moves()
                self.decide_winner()
                self.add_points()
                self.move_counter += 1
                # print(str(self.move_counter))

                if self.move_counter%4 == 0:
                    self.message += '\n' + id_to_name(self.player_names[self.current_winner]) + " has won the set! They can start the next round."
                    if self.current_winner%2 != self.tao_fen_person%2:
                        # qiang fen won 
                        self.qiang_fen_points += self.points
                        if self.points > 0:
                            self.message += "\n抢分 just won " + str(self.points) + " points! They now have " + str(self.qiang_fen_points) + " points."
                    elif self.points >= 10:
                        self.message += "\n逃分 just stole " + str(self.points) + " points! xD"
                    self.points = 0
                    self.first_player = True
                    self.first_player_id = self.current_winner
                    self.moving_player = self.first_player_id
                    self.mode = 0
                    # print(self.first_player_id)
                    self.clear_moves()
                    if not len(self.player_decks[0]):
                        self.game_ended = True 
                        self.actually_playing = False
                        return self.endround()
                    if self.is_bot[self.moving_player]:
                        move = self.bot_objs[self.first_player_id].make_move(0, 0, [])
                        self.first_player = False
                        self.moving_player = (self.moving_player)+1%4
                        self.print_moves()
                        self.decide_winner()
                        self.add_points()
                        self.move_counter += 1
                        self.message += "\nCurrently, " + id_to_name(self.player_names[self.current_winner]) + " is big."
                        self.message += '\n' + id_to_name(self.player_names[(self.moving_player+1)%4]) + ", it's your turn!"

                else:
                    self.first_player = False
                    self.moving_player = (self.moving_player+1)%4
                    if self.is_bot[self.moving_player]:
                        move = self.bot_objs[self.moving_player].make_move(len(self.player_moves[self.first_player_id]), self.mode, self.player_moves[self.first_player_id])
                        self.player_moves[self.moving_player] = move
                        self.print_moves()
                        self.decide_winner()
                        self.add_points()
                        self.move_counter += 1
                        if self.move_counter%4 == 0:
                            self.message += '\n' + id_to_name(self.player_names[self.current_winner]) + " has won the set! They can start the next round."
                            if self.current_winner%2 != self.tao_fen_person%2:
                                # qiang fen won 
                                self.qiang_fen_points += self.points
                                if self.points > 0:
                                    self.message += "\n抢分 just won " + str(self.points) + " points! They now have " + str(self.qiang_fen_points) + " points."
                            elif self.points >= 10:
                                self.message += "\n逃分 just stole " + str(self.points) + " points! xD"
                            self.points = 0
                            self.first_player = True
                            self.first_player_id = self.current_winner
                            self.moving_player = self.first_player_id
                            self.mode = 0
                            # print(self.first_player_id)
                            self.clear_moves()
                            if not len(self.player_decks[0]):
                                self.game_ended = True 
                                self.actually_playing = False
                                return self.endround()

                    self.message += "\nCurrently, " + id_to_name(self.player_names[self.current_winner]) + " is big."
                    self.message += '\n' + id_to_name(self.player_names[self.moving_player%4]) + ", it's your turn!"

                # PM next person 
                self.print_deck(self.player_names[self.moving_player])

                self.to_confirm = False
                self.confirming = -1 
                return self.message
            else:
                # not valid 
                if self.first_player:
                    self.mode = 0
                for i in range(len(self.player_moves[self.moving_player])):
                    self.player_decks[self.moving_player].append(self.player_moves[self.moving_player][i])
                # self.print_deck(self.player_names[self.moving_player])
                self.player_moves[self.moving_player].clear()
                return self.personal_message

        # end round, start next round! 
        elif self.game_ended:
            print("Game ended!")
            # self.actually_playing = False
            self.reset_vars()
            # self.game_ended = False 
            return self.message

    def endround(self):
        if self.current_winner%2 != self.tao_fen_person:
            # qiang fen won the last set, must add di points
            if self.di_points > 0:
                self.qiang_fen_points += self.di_points
                self.message = "Since " + id_to_name(self.current_winner) + " won the last set, they get " + str(self.di_points) + " more points!"

        self.message += "\nIn total, 抢分 has won " + str(self.qiang_fen_points) + " points!"

        if self.qiang_fen_points < 20:
            self.message += "\nThis means that ~~they kinda suck~~ 逃分 gets to skip a trump, so the trump will go from " + self.trump_value
            self.message += " to" + self.deck_nums[self.deck_nums.index(self.trump_value)+2] + '. Congrats!!!!!!!'
            # self.message += '\nThe first player will change from ' + id_to_name(self.player_names[self.tao_fen_person]) + " to " + id_to_name(self.player_names[(self.tao_fen_person+2)%4]) + "!!"
            self.tao_fen_person =(self.tao_fen_person+2)%4
            self.trump_value = self.deck_nums[self.deck_nums.index(self.trump_value)+2]

        elif self.qiang_fen_points < 80:
            self.message += "\nThis means that 逃分 _ascends_ to the next trump, so the trump will go from " + self.trump_value
            self.message += " to" + self.deck_nums[self.deck_nums.index(self.trump_value)+1] + '. Congrats!!!!!!!'
            # self.message += '\nThe first player will change from ' + id_to_name(self.player_names[self.tao_fen_person]) + " to " + id_to_name(self.player_names[(self.tao_fen_person+2)%4]) + "!!"
            self.tao_fen_person = (self.tao_fen_person+2)%4
            self.trump_value = self.deck_nums[self.deck_nums.index(self.trump_value)+1]

        elif self.qiang_fen_points < 120:
            # change roles!
            self.message += "\n抢分 wins! Wow! So amazing!!! So they get to be 逃分 now and the trump will be " + self.other_trump + '.'
            self.trump_value, self.other_trump = self.other_trump, self.trump_value
            self.tao_fen_person = (self.tao_fen_person+1)%4

        else:
            # change roles!
            self.message += "\n抢分 wins!!! By a lot!!!!!! Yay!! So they get to be 逃分 now and the trump will be "
            self.other_trump = self.deck_nums[self.deck_nums.index(self.other_trump)+1]
            self.message += self.other_trump + '.'
            self.trump_value, self.other_trump = self.other_trump, self.trump_value
            self.tao_fen_person = (self.tao_fen_person+1)%4

        print("Next tao fen person: " + str(self.tao_fen_person))

        self.update_score()
        self.message += "\n" + id_to_name(self.player_names[self.tao_fen_person]) + ", click /confirm to start the next round. (cuz ur drawing the 底)"
        self.confirming = self.tao_fen_person
        self.to_confirm = True
        return self.message

    def reset_vars(self):
        self.deck = [] 
        self.draw_counter = 0
        self.deck_counter = 0
        self.player_decks = {0:[], 1:[], 2:[], 3:[]}

        self.decided_trump = False
        self.one_trumps_played = False
        self.two_trumps_played = False 

        self.making_di = False
        self.di = []
        self.di_points = 0

        self.to_confirm = False
        self.confirming = -1
        
        self.actually_playing = False
        self.moving_player = -1
        self.first_player = True
        self.first_player_id = self.tao_fen_person
        self.player_moves = {0:[], 1:[], 2:[], 3:[]}
        self.mode = 0
        self.move_counter = 0
        self.aces_suits = []
        self.current_winner = -1
        self.same_pattern = False 
        self.qiang_fen_points = 0
        # self.game_ended = False

        self.make_deck()
        self.auto_deal()

    def other(self, card, name):
        if not self.tao_fen:
            return self.check_tao_fen(card, name)

        elif not self.decided_trump:
            return self.check_trump_suit(card, name)

        elif self.making_di:
            return self.check_making_di(card, name)

        elif self.actually_playing and self.player_names.index(name) == self.moving_player:
            return self.check_playing_cards(card, name)

        else:
            return ''

    def whos_big(self):
        if self.current_winner == -1:
            return "No one yet."
        return id_to_name(self.player_names[self.current_winner])

    def replace(self, name1, name2):
        if name1 in self.player_names:
            index = self.player_names.index(name1)
            self.player_names[index] = name2
            index = tractor_selfids.index(name1)
            tractor_selfids[index] = name2
            return "Success! " + id_to_name(name2) + " is now in the game."
        else:
            index = self.player_names.index(name2)
            self.player_names[index] = name1
            index = tractor_selfids.index(name2)
            tractor_selfids[index] = name1
            return "Success! " + id_to_name(name1) + " is now in the game."

    def whoseturn(self):
        return id_to_name(self.player_names[self.moving_player])

    def update_score(self):
        global tractor_scores, tractor_standings, tractor_rawscores
        winner1 = id_to_name(self.player_names[self.tao_fen_person])
        winner2 = id_to_name(self.player_names[(self.tao_fen_person+2)%4])
        loser1 = id_to_name(self.player_names[(self.tao_fen_person+1)%4])
        loser2 = id_to_name(self.player_names[(self.tao_fen_person+3)%4])
        number_of_entries = tractor_rawscores['E1']
        new_entry = 'A' + str(number_of_entries.value+2)
        tractor_rawscores[new_entry].value = datetime.today()
        new_entry = 'B' + str(number_of_entries.value+2)
        tractor_rawscores[new_entry].value = winner1
        new_entry = 'C' + str(number_of_entries.value+2)
        tractor_rawscores[new_entry].value = winner2
        number_of_entries.value += 1

        number_of_winners = tractor_standings['E1']
        one_found = False
        two_found = False
        three_found = False
        four_found = False
        for i in range(2, number_of_winners.value+2):
            cell = 'A' + str(i)
            name = tractor_standings[cell].value
            if name == winner1 or name == winner2 or name == loser1 or name == loser2:
                if name == winner1:
                    one_found = True
                elif name == winner2:
                    two_found = True
                elif name == loser1:
                    three_found = True
                else:
                    four_found = True
                if name == winner1 or name == winner2:
                    cell = 'B' + str(i)
                    tractor_standings[cell].value += 1
                cell = 'C' + str(i)
                tractor_standings[cell].value += 1

        if not one_found:
            number_of_winners.value += 1
            cell = 'A' + str(number_of_winners.value)
            tractor_standings[cell].value = winner1
            cell = 'B' + str(number_of_winners.value)
            tractor_standings[cell].value = 1
            cell = 'C' + str(number_of_winners.value)
            tractor_standings[cell].value = 1

        if not two_found:
            number_of_winners.value += 1
            cell = 'A' + str(number_of_winners.value)
            tractor_standings[cell].value = winner2
            cell = 'B' + str(number_of_winners.value)
            tractor_standings[cell].value = 1
            cell = 'C' + str(number_of_winners.value)
            tractor_standings[cell].value = 1

        if not three_found:
            number_of_winners.value += 1
            cell = 'A' + str(number_of_winners.value)
            tractor_standings[cell].value = loser1
            cell = 'B' + str(number_of_winners.value)
            tractor_standings[cell].value = 0
            cell = 'C' + str(number_of_winners.value)
            tractor_standings[cell].value = 1

        if not four_found:
            number_of_winners.value += 1
            cell = 'A' + str(number_of_winners.value)
            tractor_standings[cell].value = loser2
            cell = 'B' + str(number_of_winners.value)
            tractor_standings[cell].value = 0
            cell = 'C' + str(number_of_winners.value)
            tractor_standings[cell].value = 1

        tractor_scores.save('80fenscores.xlsx')

    def print_stats(self):
        global tractor_scores, tractor_standings, tractor_rawscores
        self.message = ''
        number_of_winners = tractor_standings['E1']
        for i in range(2, number_of_winners.value+1):
            cell = 'A' + str(i)
            # print("Cell: " + cell)
            self.message += '\n' + tractor_standings[cell].value
            cell = 'B' + str(i)
            # print("Cell: " + cell)
            self.message += ' - ' + str(tractor_standings[cell].value)
            wins = int(tractor_standings[cell].value)
            cell = 'C' + str(i)
            played = int(tractor_standings[cell].value)
            win_rate = wins/played*100
            win_rate = round(win_rate, 1)
            self.message += ' (' + str(win_rate) + '%)'
        return self.message

class tictactoe():
  def __init__(self):
    rows, cols = (9, 9)
    self.board = [['.']*rows for _ in range(cols)]
    self.won = [[' ']*3 for _ in range(3)]
    self.turn = 'X'
    self.xcor = -1 
    self.ycor = -1 
    self.message = ''
    self.game_ended = False
    self.num_occupied = [[0]*3 for _ in range(3)]
    self.num_big = 0
    self.is_draw = False 

  def check_square(self, x, y):
    square_x = math.floor(x/3)
    square_y = math.floor(y/3)
    if square_x == self.xcor and square_y == self.ycor:
      return True
    return False 

  def check_small_win(self, x, y):
    square_x = math.floor(x/3)
    square_y = math.floor(y/3)
    # 1
    # ONN
    # ONN
    # ONN
    if self.board[square_x*3][square_y*3] == self.board[square_x*3+1][square_y*3] == self.board[square_x*3+2][square_y*3] != '.':
      return self.board[square_x*3][square_y*3]
    # 2
    # OOO
    # NNN 
    # NNN 
    if self.board[square_x*3][square_y*3] == self.board[square_x*3][square_y*3+1] == self.board[square_x*3][square_y*3+2] != '.':
      return self.board[square_x*3][square_y*3]
    # 3
    # ONN
    # NON
    # NNO
    if self.board[square_x*3][square_y*3] == self.board[square_x*3+1][square_y*3+1] == self.board[square_x*3+2][square_y*3+2] != '.':
      return self.board[square_x*3][square_y*3]
    # 4
    # NNO
    # NON 
    # ONN
    if self.board[square_x*3+2][square_y*3] == self.board[square_x*3+1][square_y*3+1] == self.board[square_x*3][square_y*3+2] != '.':
      return self.board[square_x*3+2][square_y*3]
    # 5
    # NON
    # NON 
    # NON
    if self.board[square_x*3][square_y*3+1] == self.board[square_x*3+1][square_y*3+1] == self.board[square_x*3+2][square_y*3+1] != '.':
      return self.board[square_x*3][square_y*3+1]
    # 6
    # NNO
    # NNO 
    # NNO
    if self.board[square_x*3][square_y*3+2] == self.board[square_x*3+1][square_y*3+2] == self.board[square_x*3+2][square_y*3+2] != '.':
      return self.board[square_x*3][square_y*3+2]
    # 7
    # NNN
    # OOO 
    # NNN
    if self.board[square_x*3+1][square_y*3] == self.board[square_x*3+1][square_y*3+1] == self.board[square_x*3+1][square_y*3+2] != '.':
      return self.board[square_x*3+1][square_y*3]
    # 8
    # NNN
    # NNN 
    # OOO
    if self.board[square_x*3+2][square_y*3] == self.board[square_x*3+2][square_y*3+1] == self.board[square_x*3+2][square_y*3+2] != '.':
      return self.board[square_x*3+2][square_y*3]
    return 'N'

  def check_big_win(self):
    if self.won[0][0] == self.won[1][0] == self.won[2][0] != ' ':
      return self.won[0][0]
    if self.won[0][0] == self.won[0][1] == self.won[0][2] != ' ':
      return self.won[0][0]
    if self.won[0][0] == self.won[1][1] == self.won [2][2] != ' ':
      return self.won[0][0];
    if self.won[1][0] == self.won[1][1] == self.won[1][2] != ' ':
      return self.won[1][0]
    if self.won[2][0] == self.won[2][1] == self.won[2][2] != ' ':
      return self.won[2][0]
    if self.won[0][1] == self.won[1][1] == self.won[2][1] != ' ':
      return self.won[0][1] 
    if self.won[0][2] == self.won[1][2] == self.won[2][2] != ' ':
      return self.won[2][2]
    if self.won[0][2] == self.won[1][1] == self.won[2][0] != ' ':
      return self.won[0][2]
    return 'N'

  def check_legit(self, x, y):
    self.message = '' 
    if self.xcor == -1 and self.ycor == -1:
      pass
    elif not self.board[x][y]== '.':
      self.message = 'Square has already been taken!' 
    elif (self.won[self.xcor][self.ycor] == 'X' or self.won[self.xcor][self.ycor] == 'O'):
      self.message = 'Square has already been won! (If you were supposed to put in this square, you can put in any square!)' 
    elif not self.check_square(x, y):
      self.message = "Wrong square?"
    elif self.won[self.xcor][self.ycor] == 'N':
      self.message = "Square is already filled!"
    return self.message
  
  def move(self, x, y):
    square_x = x%3
    square_y = y%3
    small_x = math.floor(x/3)
    small_y = math.floor(y/3)
    if self.check_legit(x, y) != '':
      return self.message 
    self.board[x][y] = self.turn 
    self.num_occupied[small_x][small_y] += 1
    if not self.check_small_win(x, y) == 'N':
      self.won[small_x][small_y] = self.check_small_win(x, y)
      self.message = self.won[small_x][small_y] + " has won " + str(small_x) + ", " + str(small_y) + "!"
      # self.xcor = -1 
      # self.ycor = -1 
      if not self.check_big_win() == 'N':
        self.message = self.check_big_win()
        self.message += " has won!"
        self.game_ended = True 
        return self.message
      self.num_big += 1
      if self.num_big == 9:
        self.game_ended = True 
        self.message = "It's a draw!"
        self.is_draw = True 
        return self.message 
    if self.turn == 'X':
      self.turn = 'O'
    else:
      self.turn = 'X'
    if self.message == '':
        self.message = "Move successful!"
    if self.num_occupied[small_x][small_y] == 9:
        self.won[small_x][small_y] == 'N'
    # self.anywhere should be true if 
    # 1. no space in square
    # 2. square has been won
    square_x = x%3
    square_y = y%3
    self.xcor = square_x 
    self.ycor = square_y
    if self.won[self.xcor][self.ycor] != ' ':
      self.xcor = -1 
      self.ycor = -1 
    return self.message 
    
  def print(self):
    string = ''
    string += "  "
    for i in range(1, 10):
      string += str(i)
      if i%3 == 0:
        string += " "
    string += '\n' 
    for i in range(0, 9):
      string += str(i+1) + " " 
      for j in range(0, 9):
        string += self.board[i][j]
        if (j+1)%3 == 0:
          string += ' '
      string += '\n' 
      if (i+1)%3 == 0:
        string += '\n'
    return string 
    
    # self.message = "Turn has been made!"
    # return self.message 

class TelegramBot:

    def __init__(self):
        """"
        Initializes an instance of the TelegramBot class.
        Attributes:
            chat_id:str: Chat ID of Telegram chat, used to identify which conversation outgoing messages should be send to.
            text:str: Text of Telegram chat
            first_name:str: First name of the user who sent the message
            last_name:str: Last name of the user who sent the message
        """

        self.chat_id = None
        self.text = None
        self.first_name = None
        self.last_name = None
        self.incoming_message_text = None
        self.outgoing_message_text = None 
        self.sender = None 

    def parse_webhook_data(self, data):
        global unixtime, current_id, send_id, msg_id, orig_text, times

        """
        Parses Telegram JSON request from webhook and sets fields for conditional actions
        Args:
            data:str: JSON string of data
        """ 

        if data['update_id'] == current_id:
            times = 0 
            self.incoming_message_text = "Edited message"
            return

        times = 0 
        current_id = data['update_id']
        try:
            message = data['message']
        except KeyError as e:
            self.incoming_message_text = 'Edited'
            return 

        self.chat_id = message['chat']['id']
        msg_id = self.chat_id
        self.first_name = message['from']['first_name']
        self.sender = message['from']['id']

        try:
            self.incoming_message_text = message['text'].lower()
            orig_text = message['text']
        except KeyError as e:
            self.incoming_message_text = ''
            return

        print(self.sender, end=", ")
        print(self.chat_id, end=", ")
        print(self.first_name, end=", ")
        print(current_id)

    def action(self):
        """
        Conditional actions based on set webhook data.
        Returns:
            bool: True if the action was completed successfully else false
        """
        global intro_cca, intro_names, intro_classes, intro_id, send_id, intro_subj, intro_game, bridge_score, bridge_name, ttt, starter, opponent, turn, orig_text, things, words, to_approve, tractor_games, tractor_ids, tractor_selfids, _wb, _results, _date, bridge_stats, standings, raw_scores

        success = None

        if self.incoming_message_text == '':
            success = True 
            return success

        # elif self.incoming_message_text.startswith('/full80fenstats'):
        #     filename = "80fenscores.xlsx"
        #     sendDocument(self.chat_id, filename)

        elif self.incoming_message_text.startswith('/addbot'):
            the_game = tractor_games[0]
            if len(the_game.player_names) == 4:
                self.outgoing_message_text = "There are already 4 people in the game!"
                success = self.send_message()
                return success 
            the_game.add_bot()
            self.outgoing_message_text = "Done!"
            if len(the_game.player_names) == 4:
                for i, name in enumerate(the_game.player_names):
                    if not the_game.is_bot[i]:
                        self.chat_id = the_game.player_names[i]
                        the_game.print_deck(the_game.player_names[i])
                        self.outgoing_message_text = the_game.personal_message
            success = self.send_message()
            return success

        elif self.incoming_message_text.startswith('/80fenstats'):
            the_game = tractor_games[0]
            self.outgoing_message_text = the_game.print_stats()
            success = self.send_message()
            return success

        elif self.incoming_message_text.startswith('/howmanypoints'):
            if self.chat_id > 0:
                self.outgoing_message_text = "Please use this command in a group chat."
                success = self.send_message()
                return success
            index = tractor_ids.index(self.chat_id)
            the_game = tractor_games[index]
            self.outgoing_message_text = "抢分 has " + str(the_game.qiang_fen_points) + " points."
            success = self.send_message()
            return success

        elif self.incoming_message_text.startswith('/whoseturn'):
            if self.chat_id > 0:
                self.outgoing_message_text = "Please use this command in a group chat."
                success = self.send_message()
                return success
            index = tractor_ids.index(self.chat_id)
            the_game = tractor_games[index]
            self.outgoing_message_text = "It is " + the_game.whoseturn() + "'s turn."
            success = self.send_message()
            return success

        elif self.incoming_message_text.startswith('/replace'):
            if self.chat_id > 0:
                self.outgoing_message_text = "This command can only be used in group chats!"
                success = self.outgoing_message_text
                return success
            name = self.incoming_message_text[9:]
            name1, name2 = name.split(", ")

            if not name1 or not name2:
                self.outgoing_message_text = "Person(s) not found! (Remember to separate using comma?) Format: /replace <name1>, <name2>"
                success = self.send_message()
                return success 

            if name1 in tractor_selfids and name2 in tractor_selfids:
                self.outgoing_message_text = "Both players are already in the game!"
                success = self.send_message()
                return success

            if self.chat_id not in tractor_ids:
                self.outgoing_message_text = "/start80fen to start a game!"
                success = self.send_message()
                return success

            the_game = tractor_games[tractor_ids.index(self.chat_id)]
            self.outgoing_message_text = the_game.replace(name1, name2)
            success = self.send_message()
            return success

        elif self.incoming_message_text.startswith('/whosbig'):
            if self.chat_id > 0:
                self.outgoing_message_text = "Please use this command in a group chat."
                success = self.send_message()
                return success
            index = tractor_ids.index(self.chat_id)
            the_game = tractor_games[index]
            self.outgoing_message_text = the_game.whos_big()
            success = self.send_message()
            return success

        elif self.incoming_message_text.startswith('/whatstrump'):
            if self.chat_id > 0:
                self.outgoing_message_text = "Please use this command in a group chat."
                success = self.send_message()
                return success
            index = tractor_ids.index(self.chat_id)
            the_game = tractor_games[index]
            self.outgoing_message_text = the_game.whats_trump()
            success = self.send_message()
            return success

        elif self.incoming_message_text.startswith('/start80fen'):
            if self.chat_id > 0: 
                self.outgoing_message_text = "This game can only be played in group chats."
                success = self.send_message()
                return success
            if self.chat_id in tractor_ids:
                self.outgoing_message_text = "There is already an ongoing game!"
                success = self.send_message()
                return success
            tractor_ids.append(self.chat_id)
            the_game = tractor()
            the_game.start()
            self.outgoing_message_text = the_game.message
            tractor_games.append(the_game)
            success = self.send_message()
            return success 

        elif self.incoming_message_text.startswith('/join80fen'):
            if self.chat_id > 0:
                self.outgoing_message_text = "This game can only be played in group chats."
                success = self.send_message()
                return success
            if not self.chat_id in tractor_ids:
                self.outgoing_message_text = "/start80fen to start a new game!"
                success = self.send_message()
                return success
            if self.sender in tractor_selfids:
                self.outgoing_message_text = "You're already in a game!"
                success = self.send_message()
                return success 

            orig_chat_id = self.chat_id
            index = tractor_ids.index(self.chat_id)
            the_game = tractor_games[index]
            if len(the_game.player_names) == 4:
                self.outgoing_message_text = "There are already 4 people in the game!"
                success = self.send_message()
                return success 

            tractor_selfids.append(self.sender)
            the_game.join(self.sender)
            if len(the_game.player_names) == 4:
                for i in range(4):
                    if not the_game.is_bot[i]:
                        self.chat_id = the_game.player_names[i]
                        the_game.print_deck(the_game.player_names[i])
                        self.outgoing_message_text = the_game.personal_message
                        success = self.send_message()
            self.chat_id = orig_chat_id
            self.outgoing_message_text = the_game.message
            success = self.send_message()
            return success

        elif self.incoming_message_text.startswith('/drawcard'):
            if self.chat_id > 0:
                self.outgoing_message_text = "This command can only be used in group chats."
                success = self.send_message()
                return success
            if not self.chat_id in tractor_ids:
                self.outgoing_message_text = "/start80fen to start a new game!"
                success = self.send_message()
                return success
            index = tractor_ids.index(self.chat_id)
            the_game = tractor_games[index]
            if not self.sender in the_game.player_names and len(the_game.player_names) == 4:
                self.outgoing_message_text = "You're not in the game!"
                success = self.send_message()
                return success
            elif len(the_game.player_names) < 4:
                self.outgoing_message_text = "Use /join80fen to join the game first!"
                success = self.send_message()
                return success 
            self.outgoing_message_text = the_game.draw(self.sender)
            if self.outgoing_message_text == the_game.personal_message:
                self.outgoing_message_text = the_game.message
                success = self.send_message()
                self.chat_id = self.sender
                self.outgoing_message_text = the_game.personal_message
                success = self.send_message()
                return 
            success = self.send_message()
            return success 

        elif self.incoming_message_text.startswith('/drawdi'):
            if not self.chat_id in tractor_ids:
                self.outgoing_message_text = "/start80fen to start a new game!"
                success = self.send_message()
                return success
            index = tractor_ids.index(self.chat_id)
            the_game = tractor_games[index]
            if not self.sender in the_game.player_names and len(the_game.player_names) == 4:
                self.outgoing_message_text = "You're not in the game!"
                success = self.send_message()
                return success
            elif len(the_game.player_names) < 4:
                self.outgoing_message_text = "Use /join80fen to join the game first!"
                success = self.send_message()
                return success 

            self.outgoing_message_text = the_game.draw_di(self.sender)
            if self.outgoing_message_text == the_game.personal_message:
                self.chat_id = self.sender
            success = self.send_message()
            return success

        elif self.incoming_message_text.startswith('/showdeck'):
            if not self.sender in tractor_selfids:
                self.outgoing_message_text = "You're not in a game!"
                success = self.send_message()
                return success
            index = math.floor(tractor_selfids.index(self.sender)/4)
            the_game = tractor_games[index]
            the_game.print_deck(self.sender)
            self.outgoing_message_text = the_game.personal_message
            success = self.send_message()
            return success 

        elif self.incoming_message_text.startswith('/confirm'):
            if not self.sender in tractor_selfids:
                self.outgoing_message_text = "You're not in a game, you have nothing to confirm xD"
                success = self.send_message()
                return success
            index = math.floor(tractor_selfids.index(self.sender)/4)
            the_game = tractor_games[index]
            self.outgoing_message_text = the_game.confirm(self.sender)
            if self.outgoing_message_text == the_game.personal_message:
                self.chat_id = self.sender
                success = self.send_message()
                return success
            elif self.outgoing_message_text == the_game.message:
                self.chat_id = tractor_ids[index]
                success = self.send_message()
                if the_game.actually_playing:
                    self.chat_id = the_game.player_names[the_game.moving_player]
                    self.outgoing_message_text = the_game.personal_message
                    success = self.send_message()
                elif self.outgoing_message_text == 'Done dealing!':
                    for i in range(4):
                        self.chat_id = the_game.player_names[i]
                        the_game.print_deck(the_game.player_names[i])
                        self.outgoing_message_text = the_game.personal_message
                        success = self.send_message()
                    the_game.game_ended = False
                return success

        elif self.incoming_message_text.startswith('/approve'):
            if not self.sender == admin_id:
                self.outgoing_message_text = "Admin only!"
                success = self.send_message()
                return success
            word = self.incoming_message_text[9:]
            if not word in to_approve:
                self.outgoing_message_text = "Word not in approval list!"
                success = self.send_message()
                return success
            words.append(word)
            words.sort()
            self.outgoing_message_text = "Word has been approved!"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/append'):
            word = self.incoming_message_text[8:]
            if not word.startswith('s'):
                self.outgoing_message_text = "Word does not start with s!"
                success = self.send_message()
                return success 
            if len(word) < 10:
                self.outgoing_message_text = "Word is too short!"
                success = self.send_message()
                return success
            if word in words:
                self.outgoing_message_text = "Word is already in list!"
                success = self.send_message()
                return success 
            if not word[len(word)-1] == 's':
                self.outgoing_message_text = "Word does not end in s!"
                success = self.send_message()
                return success
            too_lazy = self.chat_id  
            self.chat_id = admin_id 
            self.outgoing_message_text = "New word to /approve : " + word 
            success = self.send_message()
            to_approve.append(word)
            self.chat_id = too_lazy 
            self.outgoing_message_text = "Word has been sent for approval!"
            success = self.send_message()
            return success 

        elif self.incoming_message_text.startswith('/schain'):
            clause = self.incoming_message_text[8:]
            print(clause)
            if clause == "" or clause == "wwcoolbot":
                self.outgoing_message_text = words[0]
                for i in range(1, len(words)):
                    self.outgoing_message_text += '\n'
                    self.outgoing_message_text += words[i] 
                success = self.send_message()
                return success
            started = False 
            self.outgoing_message_text = ""
            for i in range(0, len(words)):
                if words[i].startswith(clause):
                    started = True 
                if started and not words[i].startswith(clause):
                    break
                if started:
                    self.outgoing_message_text += words[i]
                    self.outgoing_message_text += '\n'
            success = self.send_message()
            return success
            
        elif self.incoming_message_text.startswith('/tinder'):
            self.outgoing_message_text = "Will be revamped soon. soz"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/rate'):
            self.outgoing_message_text = "Will be revamped soon. soz"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/soulmate'):
            self.outgoing_message_text = "Will be revamped soon. soz"
            success = self.send_message()
            
        elif self.incoming_message_text.startswith('/commands'):
            self.outgoing_message_text = "Hi there. These are the following commands I understand:\n\n*Introductions*\n/intro - set your intro (pms only)\n/getintro <name> - get intro of a person\n\n*Bridge*\n/reportbridge <winner1>, <winner2> - report bridge winners\n/statsbridge - see how many games each person has won\n\n*Tictactoe*\n/newttt <name> - create new tic tac toe game with someone\n/move <vert, hor> - make move for tic tac toe\n/printboard <name> - print someone's tic tac toe gameboard\n/resign - resign current tic tac toe game\n\n*Fake tinder (Very beta, do not recommend)*\n/tinder - start rating!\n/rate <num> - rate person/thing\n/soulmate - get your soulmate!\n\n*Others*\n/games - current list of games played in chat and rules on how to play them\n/hug <name> - give someone a hug!\n/kiss <name> - oh the horror"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/resign'):
            if not (self.sender in starter or self.sender in opponent):
                self.outgoing_message_text = "You're resigning for no reason? You haven't started a game yet."
                success = self.send_message()
                return success 
            # starter = None 
            if self.sender in starter:
                index = starter.index(self.sender)
                winner = id_to_name(opponent[index])
            else:
                index = opponent.index(self.sender)
                winner = id_to_name(starter[index])
            self.outgoing_message_text = "{} has won! Use /newttt <name> to start a new game.".format(winner)
            success = self.send_message()
            self.outgoing_message_text = "{} is an amazing player/genius.".format(winner)
            success = self.send_message()
            ttt.pop(index)
            starter.pop(index)
            opponent.pop(index)
            turn.pop(index)

        elif self.incoming_message_text.startswith('/newttt'):
            opponent_id = name_to_id(self.incoming_message_text[8:])
            sender_id = self.sender
            if opponent_id == 0:
                self.outgoing_message_text = "Person not found. (format: /newttt <name>)"
                success = self.send_message()
                return success 
            if sender_id in starter or sender_id in opponent:
                self.outgoing_message_text = "You already have an ongoing game!"
                success = self.send_message()
                return success 
            if opponent_id in starter or opponent_id in opponent:
                self.outgoing_message_text = opponent + " already has an ongoing game!"
                success = self.send_message()
                return success 
            new = tictactoe()
            ttt.insert(0, new)
            starter.insert(0, sender_id)
            opponent.insert(0, opponent_id)
            turn.insert(0, sender_id)
            self.outgoing_message_text = '```'
            self.outgoing_message_text += ttt[0].print()
            self.outgoing_message_text += '```'
            success = self.send_message()
            self.outgoing_message_text = "Game successfully created! Use /move <x, y> to make your first move :cD"
            success = self.send_message()
            # self.outgoing_message_text = "{} just started a tic tac toe game with you! (If you are playing in a group chat, use /turnoff or /turnon to stop receiving notifs or receive them)"
            # self.chat_id = opponent_id 
            # success = self.send_message()

        elif self.incoming_message_text.startswith('/move'):
            coords = self.incoming_message_text[6:]
            sender_id = self.sender 
            if not sender_id in turn:
                self.outgoing_message_text = "It's not your turn yet. Or maybe you haven't started a game. I don't know. Hehe."
                success = self.send_message()
                return success 

            game = turn.index(sender_id)

            xcor = '' 
            ycor = '' 
            is_x = True 
            # print(coords)
            for i in range(0, len(coords)):
                if coords[i] == ',':
                    is_x = False 
                elif coords[i] == ' ':
                    continue 
                elif is_x:
                    xcor += coords[i] 
                else:
                    ycor += coords[i] 
            # print(xcor, ycor)
            xcor = int(xcor)
            ycor = int(ycor)
            if xcor <= 0 or xcor > 9 or ycor <= 0 or ycor > 9: 
                self.outgoing_message_text = "Invalid coordinates!"
                success = self.send_message()
                return success 

            self.outgoing_message_text = ttt[game].move(xcor-1, ycor-1)
            success = self.send_message()
            if self.outgoing_message_text == "Wrong square?" or self.outgoing_message_text == 'Square has already been won! (If you were supposed to put in this square, you can put in any square!)' or self.outgoing_message_text == "Square has already been taken!":
                return success 
            if ttt[game].game_ended and not ttt[game].is_draw:
                if self.sender in starter: # x won, starter won 
                    winner = id_to_name(starter[game])
                else:
                    winner = id_to_name(opponent[game])
                self.outgoing_message_text = "The game has ended! {} has won. Use /newttt <name> to start a new game!".format(winner)
                success = self.send_message()
                self.outgoing_message_text = "{} is an amazing player/genius.".format(winner)
                success = self.send_message()
                ttt.pop(game)
                starter.pop(game)
                opponent.pop(game)
                turn.pop(game)
                return success 
            elif ttt[game].game_ended:
                self.outgoing_message_text += ' Use /newttt <name> to start a new game!'
                success = self.send_message()
                ttt.pop(game)
                starter.pop(game)
                opponent.pop(game)
                turn.pop(game)
                return success 
            self.outgoing_message_text = "Next move must be in: "
            self.outgoing_message_text += str(ttt[game].xcor+1) + ", " + str(ttt[game].ycor+1)
            # success = self.send_message()

            if turn[game] == starter[game]:
                turn[game] = opponent[game]
            else:
                turn[game] = starter[game]
            self.outgoing_message_text += "\nIt is now " + id_to_name(turn[game]) + "\'s turn!"
            # success = self.send_message()
            self.outgoing_message_text += '\n```'
            self.outgoing_message_text += ttt[game].print()
            self.outgoing_message_text += '```'
            success = self.send_message()

        elif self.incoming_message_text.startswith('/printboard'):
            name_game = self.incoming_message_text[12:]
            name_game_id = name_to_id(name_game)
            if name_game_id == 0:
                self.outgoing_message_text = "Person not found! (Format: /printboard <name>)"
                success = self.send_message()
                return success 
            if not (name_game_id in starter or name_game_id in opponent):
                self.outgoing_message_text = "{} doesn't have an ongoing game.".format(name_game)
                success = self.send_message()
                return success 
            if name_game_id in starter:
                game = starter.index(name_game_id)
            else:
                game = opponent.index(name_game_id)
            self.outgoing_message_text = "Here is the board for the match between {} and {}".format(id_to_name(starter[game]), id_to_name(opponent[game]))
            success = self.send_message()
            self.outgoing_message_text = '```'
            self.outgoing_message_text += ttt[game].print()
            self.outgoing_message_text += '```'
            success = self.send_message()
            move_per = turn[game]
            if move_per == self.sender:
                name = "your"
            else:
                name = id_to_name(move_per) + "\'s"
            self.outgoing_message_text = "It is now {} turn, and the move must be in square {}, {}.".format(name, str(ttt[game].xcor+1), str(ttt[game].ycor+1))
            success = self.send_message()

        elif self.incoming_message_text.startswith('/statsbridge'):
            number_of_winners = standings['E1']
            self.outgoing_message_text = ''
            for i in range(1, number_of_winners.value+1):
                cell = 'A' + str(i)
                print("Cell: " + cell)
                self.outgoing_message_text += '\n' + standings[cell].value
                cell = 'B' + str(i)
                print("Cell: " + cell)
                self.outgoing_message_text += ' - ' + str(standings[cell].value)
            success = self.send_message()

        elif self.incoming_message_text.startswith('/report'):
            name = self.incoming_message_text[8:]
            name1, name2 = name.split(', ')

            name1 = id_to_name(name_to_id(name1))
            name2 = id_to_name(name_to_id(name2))

            if name1 == '' or name2 == '':
                self.outgoing_message_text = "Person(s) not found! (Remember to put a comma?)"
                success = self.send_message()
                return success 

            if name1 == name2:
                self.outgoing_message_text = "Tsk stop trying to cheat :c("
                success = self.send_message()
                return success

            update_bridge(name1, name2)

            self.outgoing_message_text = "Update successful!"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/reset'):
            for i in bridge_score:
                index = bridge_score.index(i)
                bridge_score[index] = 0
            self.outgoing_message_text = "Scores reset!"
            success = self.send_message()
        
        # fixed 6/4/2021
        elif self.incoming_message_text == '/start' or self.incoming_message_text == '/start@wwcoolbot':
            self.outgoing_message_text = "This is for managing ww cool. Use /commands to see the full list of commands."
            success = self.send_message()

        elif self.incoming_message_text.startswith('/intro'):
            if self.chat_id < 0:
                self.outgoing_message_text = "Please PM me."
                success = self.send_message()
                return success 
            if not self.chat_id in intro_id:
                intro_id.insert(0, self.chat_id)
                intro_names.insert(0, 'nil')
                intro_classes.insert(0, 'nil')
                intro_subj.insert(0, 'nil')
                intro_cca.insert(0, 'nil')
                intro_game.insert(0, 'nil')
                self.outgoing_message_text = "What is your name? (use /name in front of your answer because the coder of this is lazy 😔)"
                success = self.send_message()
            else:
                self.outgoing_message_text = "If you haven't finished, pls go scroll to find where u left off bc i lazy"
                success = self.send_message() 
            # intro_names.insert(intros, self.first_name)

        elif self.incoming_message_text.startswith('/name'):
            if self.chat_id < 0:
                self.outgoing_message_text = "Please PM me."
                success = self.send_message()
                return success
            if not self.chat_id in intro_id:
                self.outgoing_message_text = "Please type /intro to begin writing your introduction :c)"
                success = self.send_message()
                return success
            nname = orig_text[6:]
            req_index = intro_id.index(self.chat_id)
            intro_names[req_index] = nname 
            self.outgoing_message_text = "What is your class? (use /class in front of answer pls pls thank you) (also ignore if finished)"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/class'):
            if self.chat_id < 0:
                self.outgoing_message_text = "Please PM me."
                success = self.send_message()
                return success
            if not self.chat_id in intro_id:
                self.outgoing_message_text = "Please type /intro to begin writing your introduction :c)"
                success = self.send_message()
                return success 
            p_class = orig_text[7:]
            req_index = intro_id.index(self.chat_id)
            intro_classes[req_index] = p_class 
            self.outgoing_message_text = "What CCA are you in? (use /cca in front of answer bc Lazy thank you) (Ignore if finished)"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/cca'):
            if self.chat_id < 0:
                self.outgoing_message_text = "Please PM me."
                success = self.send_message()
                return success
            if not self.chat_id in intro_id:
                self.outgoing_message_text = "Please type /intro to begin writing your introduction :c)"
                success = self.send_message()
                return success
            req_index = intro_id.index(self.chat_id)
            cca = orig_text[5:]
            intro_cca[req_index] = cca 
            self.outgoing_message_text = "What's your subject combi? (use /subj in front of ans woo)"
            success = self.send_message()
            
        elif self.incoming_message_text.startswith('/subj'):
            if self.chat_id < 0:
                self.outgoing_message_text = "Please PM me."
                success = self.send_message()
                return success
            if not self.chat_id in intro_id:
                self.outgoing_message_text = "Please type /intro to begin writing your introduction :c)"
                success = self.send_message()
                return success
            index = intro_id.index(self.chat_id)
            subj_combi = orig_text[6:]
            intro_subj[index] = subj_combi
            self.outgoing_message_text = "Pick a game: WW, bridge, wordchain, or quizarium? (Use /game before answer!)"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/game') and not self.incoming_message_text.startswith('/games'):
            if self.chat_id < 0:
                self.outgoing_message_text = "Please PM me."
                success = self.send_message()
                return success
            if not self.chat_id in intro_id:
                self.outgoing_message_text = "Please type /intro to begin writing your introduction :c)"
                success = self.send_message()
                return success
            index = intro_id.index(self.chat_id)
            game = orig_text[6:]
            intro_game[index] = game 
            self.outgoing_message_text = "Intro done! Use /getintro <name> to get introduction of person"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/getintro'):
            req_name = self.incoming_message_text[10:]
            if req_name == 'all':
                for i in intro_id:
                    index = intro_id.index(i)
                    self.outgoing_message_text = "Name: "
                    self.outgoing_message_text += intro_names[index]
                    # self.outgoing_message_text += ' tg://user?id='+str(intro_id[index])
                    # self.outgoing_message_text += ' _testing_'
                    self.outgoing_message_text += "\nClass: "
                    self.outgoing_message_text += intro_classes[index]
                    self.outgoing_message_text += "\nCCA: "
                    self.outgoing_message_text += intro_cca[index]
                    self.outgoing_message_text += "\nSubject combi: "
                    self.outgoing_message_text += intro_subj[index]
                    self.outgoing_message_text += "\nFavourite game: "
                    self.outgoing_message_text += intro_game[index]
                    success = self.send_message()
                return success 
            req_id = name_to_id(req_name)
            if req_id == 0:
                for i in intro_names:
                    if i in req_name:
                        index = intro_id.index(i)
                        req_id = intro_id[index]
            if not req_id in intro_id:
                self.outgoing_message_text = "Person/intro not found."
                success = self.send_message()
                return success 
            index = intro_id.index(req_id)
            self.outgoing_message_text = "Name: "
            self.outgoing_message_text += intro_names[index]
            # self.outgoing_message_text += ' tg://user?id='+str(intro_id[index])
            # self.outgoing_message_text += ' @eggieandree test'
            self.outgoing_message_text += "\nClass: "
            self.outgoing_message_text += intro_classes[index]
            self.outgoing_message_text += "\nCCA: "
            self.outgoing_message_text += intro_cca[index]
            self.outgoing_message_text += "\nSubject combi: "
            self.outgoing_message_text += intro_subj[index]
            self.outgoing_message_text += "\nFavourite game: "
            self.outgoing_message_text += intro_game[index]
            success = self.send_message()

        elif self.incoming_message_text.startswith('/games'):
            self.outgoing_message_text = "Games:\n- Werewolf (/ww)\n- Bridge (/bridge)\n- Quizarium (/quizarium)\n- Word chain (/wordchain)"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/ww'):
            self.outgoing_message_text = "ww rules:\nIf you're good, find the bad guys. If you're bad, kill everyone else. If you're neutral, do as your role dictates. /rolelist for all roles."
            success = self.send_message()

        elif self.incoming_message_text.startswith('/bridge'):
            self.outgoing_message_text = "Bridge rules: https://tinyurl.com/wwcoolbridgerules"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/quizarium'):
            self.outgoing_message_text = "Quizarium rules:\n- Answer the questions.\nPlease don't change the language, thanks :c)"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/wordchain'):
            self.outgoing_message_text = "Word chain!! Chain words."
            success = self.send_message()

        elif self.incoming_message_text.startswith('/tictactoe'):
            self.outgoing_message_text = "Tic tac toe in tic tac toe. Win in big tic tac toe by winning in small tic tac toe. Next move is decided by position of previous move. Eg, if previous move was in bottom right of small square, next move must be in bottom right of large square."
            success = self.send_message()

        elif self.incoming_message_text.startswith('/id'):
            if not self.chat_id == admin_id: 
                self.outgoing_message_text = "Admin only!"
                success = self.send_message()
                return success 
            send_id = name_to_id(self.incoming_message_text[4:])
            self.outgoing_message_text = "ID set!"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/send'):
            if not self.chat_id == admin_id:
                self.outgoing_message_text = "Admin only!"
                success = self.send_message()
                return success 
            if send_id == 0:
                self.outgoing_message_text = "No ID set!"
                success = self.send_message()
                return success 
            mmessage = orig_text[6:]
            self.outgoing_message_text = mmessage 
            self.chat_id = send_id
            success = self.send_message()
            self.chat_id = admin_id
            self.outgoing_message_text = "Message sent!"
            success = self.send_message()

        elif self.incoming_message_text.startswith('/update'):
            if not self.chat_id == admin_id:
                self.outgoing_message_text = "Admin only!"
                success = self.send_message()
                return success 
            self.outgoing_message_text = ""
            for i in intro_id:
                self.chat_id = i
                success = self.send_message()

        elif self.incoming_message_text.startswith('/hug'):
            req_id = name_to_id(self.incoming_message_text[5:])
            if req_id == 0:
                self.outgoing_message_text = "Person not found."
                success = self.send_message()
                return success 

            sender = id_to_name(self.sender)
            if sender == '':
                sender = self.first_name

            # self.chat_id = req_id 
            self.outgoing_message_text = sender + " gave you a hug!"
            self.chat_id = req_id
            success = self.send_message()

        elif self.incoming_message_text.startswith('/') and len(self.incoming_message_text) < 5 and len(tractor_games) > 0:
            if self.chat_id < 0:
                self.outgoing_message_text = "Please PM me."
                success = self.send_message()
                return success
            if not self.sender in tractor_selfids:
                self.outgoing_message_text = "You're not in a game."
                success = self.send_message()
                return success 
            index = math.floor(tractor_selfids.index(self.sender)/4)
            the_game = tractor_games[index]
            card = orig_text[1:]
            self.outgoing_message_text = the_game.other(card, self.sender)
            if not self.outgoing_message_text == '':
                success = self.send_message()
                return success 
            return 

        return success

    def send_message(self):
        """
        Sends message to Telegram servers.
        """
        global parseMode

        res = requests.get(TELEGRAM_SEND_MESSAGE_URL.format(self.chat_id, self.outgoing_message_text))

        return True if res.status_code == 200 else False
    

    @staticmethod
    def init_webhook(url):
        """
        Initializes the webhook
        Args:
            url:str: Provides the telegram server with a endpoint for webhook data
        """

        requests.get(url)
